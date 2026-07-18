"""
File Prep — convert a raw Excel file into a clean Niche Upload-ready Excel file.
Pattern-based field detection — no column position assumptions except gmb_col / name_col.
"""
import re
import io
from typing import Optional, Dict, Any, List, Tuple, Callable

import openpyxl
from openpyxl import Workbook

from analyzers.ollama_client import ask
from analyzers.text_cleaner import clean_review_text
from analyzers.website_email import scrape_website_contact_info
from analyzers.social_screenshot import find_email_from_socials

# ── Patterns ──────────────────────────────────────────────────────────────────

_GMB_RE          = re.compile(r'google\.com/maps', re.I)
_RATING_RE       = re.compile(r'^[1-5]\.\d$')
_REVIEWS_RE      = re.compile(r'^-?\d+\.0$|^-?\d+$')
_ADDRESS_RE      = re.compile(r'^\d+\s+\w')
_PHONE_DIGITS_RE = re.compile(r'\d')
_LAT_RE          = re.compile(r'!3d(-?\d+\.\d+)')
_LON_RE          = re.compile(r'!4d(-?\d+\.\d+)')
_OWNER_STRUCTURED = re.compile(r'\b(principal|business management|additional contact)\b', re.I)
_OWNER_WITH_CONTACT = re.compile(
    r'\b(owner|manager)\b.{0,200}\b(email|calling|call us|reach us|reach the|contact us|@)\b',
    re.I | re.S
)
_EMAIL_RE        = re.compile(r'[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-z]{2,}')

_US_STATES = {
    "Alabama","Alaska","Arizona","Arkansas","California","Colorado","Connecticut",
    "Delaware","Florida","Georgia","Hawaii","Idaho","Illinois","Indiana","Iowa",
    "Kansas","Kentucky","Louisiana","Maine","Maryland","Massachusetts","Michigan",
    "Minnesota","Mississippi","Missouri","Montana","Nebraska","Nevada",
    "New Hampshire","New Jersey","New Mexico","New York","North Carolina",
    "North Dakota","Ohio","Oklahoma","Oregon","Pennsylvania","Rhode Island",
    "South Carolina","South Dakota","Tennessee","Texas","Utah","Vermont",
    "Virginia","Washington","West Virginia","Wisconsin","Wyoming",
}

_STATE_ABBR = {
    "AL":"Alabama","AK":"Alaska","AZ":"Arizona","AR":"Arkansas","CA":"California",
    "CO":"Colorado","CT":"Connecticut","DE":"Delaware","FL":"Florida","GA":"Georgia",
    "HI":"Hawaii","ID":"Idaho","IL":"Illinois","IN":"Indiana","IA":"Iowa",
    "KS":"Kansas","KY":"Kentucky","LA":"Louisiana","ME":"Maine","MD":"Maryland",
    "MA":"Massachusetts","MI":"Michigan","MN":"Minnesota","MS":"Mississippi",
    "MO":"Missouri","MT":"Montana","NE":"Nebraska","NV":"Nevada","NH":"New Hampshire",
    "NJ":"New Jersey","NM":"New Mexico","NY":"New York","NC":"North Carolina",
    "ND":"North Dakota","OH":"Ohio","OK":"Oklahoma","OR":"Oregon","PA":"Pennsylvania",
    "RI":"Rhode Island","SC":"South Carolina","SD":"South Dakota","TN":"Tennessee",
    "TX":"Texas","UT":"Utah","VT":"Vermont","VA":"Virginia","WA":"Washington",
    "WV":"West Virginia","WI":"Wisconsin","WY":"Wyoming",
}

_STATE_FROM_ADDR_RE = re.compile(r',\s*([A-Z]{2})\s+\d{5}')


_STATE_FROM_TEXT_RE = re.compile(
    r'\b(' + '|'.join(re.escape(s) for s in _US_STATES) + r')\b'
)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _clean_phone(raw: str) -> str:
    raw = str(raw).strip()
    if raw.startswith('='):
        raw = raw.lstrip('=').strip()
    digits = re.sub(r'[^\d+]', '', raw)
    return digits if len(digits) >= 7 else raw


def _extract_lat_lon(gmb_url: str) -> Tuple[Optional[float], Optional[float]]:
    lat_m = _LAT_RE.search(gmb_url)
    lon_m = _LON_RE.search(gmb_url)
    return (float(lat_m.group(1)) if lat_m else None,
            float(lon_m.group(1)) if lon_m else None)


def _is_rating(val: str) -> bool:
    return bool(_RATING_RE.match(val.strip()))


def _is_reviews(val: str, already_has_rating: bool) -> bool:
    val = val.strip()
    if not _REVIEWS_RE.match(val):
        return False
    num = abs(float(val))
    if already_has_rating and 1.0 <= num <= 5.0 and '.' in val:
        return False
    return True


def _is_address(val: str) -> bool:
    return bool(_ADDRESS_RE.match(val.strip()))


_PLUS_CODE_RE = re.compile(r'^\d{4,6}\+[A-Z0-9]{2,4}$', re.I)

def _is_phone(val: str) -> bool:
    raw = str(val).strip()
    if raw.startswith('='):
        raw = raw.lstrip('=').strip()
    # Exclude Google Plus Codes (e.g. "73530+10052", "8FRC+GJ")
    if _PLUS_CODE_RE.match(raw):
        return False
    digits = len(_PHONE_DIGITS_RE.findall(raw))
    return 7 <= digits <= 15


def _is_owner_info(val: str) -> bool:
    if len(val.split()) < 10:
        return False
    # Structured business contact block (e.g. "Principal Contact", "Business Management")
    if _OWNER_STRUCTURED.search(val):
        return True
    # Owner/manager mentioned alongside contact info (email, phone, reach)
    if _OWNER_WITH_CONTACT.search(val):
        return True
    return False


def _ollama_extract(owner_info: str) -> Tuple[str, str]:
    prompt = f"""Extract the owner's full name and their personal/business email address from this text.
Reply in EXACTLY this format (two lines only):
OWNER_NAME: [full first and last name, or blank if not found]
EMAIL: [email address only if explicitly mentioned in the text, otherwise blank]

Text: \"\"\"{owner_info[:800]}\"\"\""""
    raw = ask(prompt, max_tokens=60, temperature=0)
    owner_name, email = "", ""
    for line in raw.splitlines():
        if line.startswith("OWNER_NAME:"):
            owner_name = line.partition(":")[2].strip()
        elif line.startswith("EMAIL:"):
            raw_email = line.partition(":")[2].strip()
            m = _EMAIL_RE.search(raw_email)
            email = m.group(0) if m else ""
    return owner_name, email


# ── Row parser ────────────────────────────────────────────────────────────────

_JUNK_CELLS = {"website", "phone", "email", "address", "name", "category",
               "closed", "open", "directions", "n/a"}

_IMAGE_EXTS  = (".png", ".jpg", ".jpeg", ".gif", ".webp", ".svg")
_SKIP_HOSTS  = ("google.com", "gstatic.com", "googleapis.com")


def _is_website(val: str) -> bool:
    if not val.startswith(("http://", "https://")):
        return False
    if any(h in val for h in _SKIP_HOSTS):
        return False
    if val.lower().endswith(_IMAGE_EXTS):
        return False
    return True


def _candidates(row_vals: List[Any], fixed_cols: set) -> List[Tuple[int, str]]:
    """Return (index, str_value) for every usable cell in the row."""
    result = []
    for i, val in enumerate(row_vals):
        if i in fixed_cols or val is None:
            continue
        s = str(val).strip()
        if not s or s.lower() in _JUNK_CELLS or s.startswith("#"):
            continue
        result.append((i, s))
    return result


def _parse_row(row_vals: List[Any], gmb_col: int, name_col: int) -> Dict[str, Any]:
    result: Dict[str, Any] = {
        "gmb_url": "", "name": "", "category": "", "address": "",
        "phone": "", "website": "", "rating": None, "reviews": None,
        "reviews_text": "", "owner_info": "", "lat": None, "lon": None,
        "state": "",
    }

    # ── Fixed columns (user-specified) ───────────────────────────────────
    result["name"]    = str(row_vals[name_col] or "").strip()
    result["gmb_url"] = str(row_vals[gmb_col]  or "").strip()
    if result["gmb_url"]:
        result["lat"], result["lon"] = _extract_lat_lon(result["gmb_url"])

    cells = _candidates(row_vals, {gmb_col, name_col})

    # ── GMB URL (if not in the designated column) ─────────────────────────
    if not result["gmb_url"]:
        for _, s in cells:
            if _GMB_RE.search(s):
                result["gmb_url"] = s
                result["lat"], result["lon"] = _extract_lat_lon(s)
                break

    # ── Rating: X.X between 1.0–5.0 ──────────────────────────────────────
    for _, s in cells:
        if _is_rating(s):
            result["rating"] = float(s)
            break

    # ── Review Count: integer number ──────────────────────────────────────
    for _, s in cells:
        if _is_reviews(s, result["rating"] is not None):
            result["reviews"] = abs(int(float(s)))
            break

    # ── Address: starts with digit + street text ──────────────────────────
    for _, s in cells:
        if _is_address(s):
            result["address"] = s
            break

    # ── Phone: 7–15 digits (handles =+1205... formula cells) ─────────────
    for _, s in cells:
        if _is_phone(s):
            result["phone"] = _clean_phone(s)
            break

    # ── Website: https:// URL, not image or Google domain ─────────────────
    for _, s in cells:
        if _is_website(s):
            result["website"] = s
            break

    # ── State: standalone state name cell ────────────────────────────────
    for _, s in cells:
        if s in _US_STATES:
            result["state"] = s
            break

    # ── Category: 1–5 words, not a phone/state/URL/rating/address/number ──
    for _, s in cells:
        if (1 <= len(s.split()) <= 5
                and not _is_phone(s)
                and not _is_rating(s)
                and not _is_address(s)
                and not _is_website(s)
                and not _is_reviews(s, True)
                and s not in _US_STATES):
            result["category"] = s
            break

    # ── Large text blocks → Owner Info + Customer Reviews ────────────────
    large_texts = [(i, s) for i, s in cells if len(s.split()) >= 10]
    review_candidates = []
    for _, s in large_texts:
        if not result["owner_info"] and _is_owner_info(s):
            result["owner_info"] = s
        else:
            review_candidates.append(s)

    if review_candidates:
        result["reviews_text"] = max(review_candidates, key=len)

    # ── State fallbacks ───────────────────────────────────────────────────
    if not result["state"] and result["address"]:
        m = _STATE_FROM_ADDR_RE.search(result["address"])
        if m:
            result["state"] = _STATE_ABBR.get(m.group(1), m.group(1))

    if not result["state"]:
        for text in [result["owner_info"], result["reviews_text"]]:
            if not text:
                continue
            m = _STATE_FROM_TEXT_RE.search(text)
            if m:
                result["state"] = m.group(1)
                break

    return result


# ── Main processor ────────────────────────────────────────────────────────────

OUTPUT_HEADERS = [
    "Business Name", "Category", "Address", "Phone", "Website",
    "Rating", "Review Count", "Customer Reviews", "Owner Name", "Email",
    "Facebook", "Instagram", "Twitter", "LinkedIn",
    "YouTube", "TikTok", "Pinterest", "Yelp",
    "Latitude", "Longitude", "State", "GMB URL",
]


def build_excel(rows: list) -> bytes:
    """Build output Excel from a list of row_result dicts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Data"
    ws.append(OUTPUT_HEADERS)
    for r in rows:
        ws.append(r["excel_row"])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def process_file(
    file_bytes: bytes,
    gmb_col: int,
    name_col: int,
    progress_callback: Optional[Callable] = None,
) -> None:
    """
    Read raw Excel and process each row.
    Calls progress_callback(current, total, message, row_result) per row.
    row_result includes excel_row — the full output row values.
    Excel is built by the caller via build_excel(job['rows']).
    """
    wb_in      = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws_in      = wb_in.active
    ws_formula = wb_formula.active

    gmb_idx  = gmb_col  - 1
    name_idx = name_col - 1

    def _cell_value(r, c):
        val = ws_in.cell(r, c).value
        # If cell is empty or an Excel error, try reading the raw formula
        # e.g. =+12055551234 shows as #ERROR! with data_only=True
        if val is None or (isinstance(val, str) and val.startswith("#")):
            formula = ws_formula.cell(r, c).value
            if formula and str(formula).startswith("="):
                return str(formula)
        return val

    rows = []
    for r in range(2, ws_in.max_row + 1):
        vals = [_cell_value(r, c) for c in range(1, ws_in.max_column + 1)]
        if any(v for v in vals):
            rows.append(vals)

    total = len(rows)

    for idx, row_vals in enumerate(rows, 1):
        parsed = _parse_row(row_vals, gmb_idx, name_idx)
        name = parsed["name"]
        if not name:
            total -= 1
            continue

        # Clean reviews_text
        if parsed["reviews_text"]:
            cleaned, _ = clean_review_text(parsed["reviews_text"])
            parsed["reviews_text"] = cleaned

        # Ollama extraction
        owner_name, ollama_email = "", ""
        if parsed["owner_info"]:
            owner_name, ollama_email = _ollama_extract(parsed["owner_info"])

        # Always scrape website — ground truth for email + socials
        site_info: Dict[str, str] = {}
        if parsed["website"]:
            try:
                site_info = scrape_website_contact_info(parsed["website"])
            except Exception:
                site_info = {}

        # Email priority: Ollama (from owner info) → Website → Facebook
        email, email_source = "", "not_found"
        if ollama_email:
            email = ollama_email
            email_source = "ollama"
        elif site_info.get("email"):
            email = site_info["email"]
            email_source = "website"

        social_cols = ["facebook", "instagram", "twitter", "linkedin", "youtube", "tiktok", "pinterest", "yelp"]
        socials_found = [p for p in social_cols if site_info.get(p)]

        # Social screenshot fallback — try Facebook → Instagram → Yelp
        screenshot_platform = ""
        if not email and socials_found:
            social_urls = {p: site_info.get(p, "") for p in social_cols}
            email, screenshot_platform = find_email_from_socials(social_urls)
            if email:
                email_source = "social"

        # Build log line
        social_note = f" · {len(socials_found)} social{'s' if len(socials_found) != 1 else ''}" if socials_found else ""
        if email_source == "ollama":
            log_line = f"{name} — {owner_name} — email via Ollama{social_note}"
            log_type = "success"
        elif email_source == "website":
            log_line = f"{name} — {owner_name or '—'} — email from website{social_note}"
            log_type = "website"
        elif email_source == "social":
            log_line = f"{name} — {owner_name or '—'} — email from {screenshot_platform} screenshot{social_note}"
            log_type = "social"
        else:
            log_line = f"{name} — no email found{social_note}"
            log_type = "warning"
            if not owner_name:
                log_line = f"{name} — owner and email not found{social_note}"
                log_type = "error"

        excel_row = [
            parsed["name"], parsed["category"], parsed["address"],
            parsed["phone"], parsed["website"], parsed["rating"],
            parsed["reviews"], parsed["reviews_text"],
            owner_name, email,
            site_info.get("facebook",  ""),
            site_info.get("instagram", ""),
            site_info.get("twitter",   ""),
            site_info.get("linkedin",  ""),
            site_info.get("youtube",   ""),
            site_info.get("tiktok",    ""),
            site_info.get("pinterest", ""),
            site_info.get("yelp",      ""),
            parsed["lat"], parsed["lon"], parsed["state"],
            parsed["gmb_url"],
        ]

        row_result = {
            "name":                name,
            "owner_name":          owner_name,
            "email":               email,
            "email_source":        email_source,
            "screenshot_platform": screenshot_platform,
            "rating":              parsed["rating"],
            "category":            parsed["category"],
            "socials":             socials_found,
            "log_line":            log_line,
            "log_type":            log_type,
            "excel_row":           excel_row,
        }

        if progress_callback:
            progress_callback(idx, total, f"Processing {idx}/{total}: {name}", row_result)
