"""
GMB Cleaner — turn a raw Google Maps scrape (CSV or Excel) into a clean file with
gmb_url, name, rating, reviews, category, address, phone, website, lat, lon.

Only gmb_url (col 0) and name (col 1) are reliably positioned. Everything after
that — including rating/reviews/category, despite the header claiming fixed
columns — shifts right whenever Google renders an extra element for that listing
(a duplicated name badge, a price-tier badge like "$20-30"), so every field from
rating onward is found by pattern instead of position, same as address/phone/website.
"""
import csv
import io
import re
from typing import List, Dict, Any, Optional, Tuple

import openpyxl
from openpyxl import Workbook

from analyzers.file_prep import _extract_lat_lon, _is_address, _is_phone, _is_website, _clean_phone

_CATEGORY_JUNK = {"·", "-", "—", "n/a", "website", "directions", "closed", "open"}
_RATING_RE = re.compile(r'^[1-5](\.\d{1,2})?$')
_REVIEWS_RE = re.compile(r'^-?\d[\d,]*$')
_PRICE_TIER_RE = re.compile(r'^\$[\d,]*(–|-)?[\d,]*$|^\${1,4}$')
_STREET_SUFFIX_RE = re.compile(
    r'\b(Rd|St|Ave|Avenue|Blvd|Boulevard|Ln|Lane|Dr|Drive|Wy|Way|Hwy|Highway|'
    r'Pl|Place|Ct|Court|Cir|Circle|Pkwy|Parkway|Route|Rte|Terrace|Ter|Trail|Loop)\.?$',
    re.I,
)


def _is_usable_category(s: str) -> bool:
    s = s.strip()
    if not s or s.lower() in _CATEGORY_JUNK or _PRICE_TIER_RE.match(s):
        return False
    return any(c.isalpha() for c in s)


def _looks_like_address(s: str) -> bool:
    """File Prep's _is_address requires a leading house number, which misses
    rural roads like "Marlboro Rd" that have none, and addresses prefixed with
    the business name or a highway name before the actual street ("Willow
    Crossing Farm, 2780 VT-15") — catch those by street suffix instead."""
    if _is_address(s):
        return True
    for part in s.split(","):
        part = part.strip()
        if _is_address(part):
            return True
    words = s.split()
    return 2 <= len(words) <= 12 and bool(_STREET_SUFFIX_RE.search(s))


def _extract_rating_reviews_category(cells: List[str], name: str) -> Tuple[Optional[float], Optional[int], str]:
    """cells = row[2:], as trimmed strings. Scans by pattern rather than trusting a
    fixed column, since a duplicated name cell or a price-tier badge ("$20-30") each
    shift every field after it one column to the right."""
    rating, rating_idx = None, -1
    for i, s in enumerate(cells):
        if s and s != name and _RATING_RE.match(s):
            rating, rating_idx = float(s), i
            break

    reviews, reviews_idx = None, rating_idx
    for i in range(rating_idx + 1, len(cells)):
        s = cells[i]
        if s and _REVIEWS_RE.match(s):
            try:
                reviews = abs(int(s.replace(",", "")))
            except ValueError:
                continue
            reviews_idx = i
            break

    category = ""
    for i in range(reviews_idx + 1, len(cells)):
        s = cells[i]
        if s and s != name and _is_usable_category(s):
            category = s
            break

    return rating, reviews, category


def _read_csv_rows(file_bytes: bytes) -> List[List[Any]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def _read_xlsx_rows(file_bytes: bytes) -> List[List[Any]]:
    """Unlike CSV, xlsx retains each cell's underlying formula alongside its
    computed value. A phone cell entered as "=+18025551234" evaluates to
    #ERROR! in Google Sheets but the formula text survives in the xlsx export —
    fall back to it wherever the computed value is missing or an error."""
    wb_val = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    wb_formula = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    ws_val = wb_val.active
    ws_formula = wb_formula.active

    rows = []
    for r in range(1, ws_val.max_row + 1):
        row = []
        for c in range(1, ws_val.max_column + 1):
            val = ws_val.cell(r, c).value
            if val is None or (isinstance(val, str) and val.strip().startswith("#")):
                formula = ws_formula.cell(r, c).value
                if formula and str(formula).startswith("="):
                    val = str(formula)
            row.append(val)
        rows.append(row)
    return rows


def _cell_str(val: Any) -> str:
    """xlsx numeric cells come back as Python int/float (e.g. -69.0), not text.
    str(-69.0) == "-69.0", which fails the reviews pattern (no decimals allowed,
    to avoid matching a comma-only cell) — normalise whole-number floats first."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


def _parse_row(row: List[Any]) -> Optional[Dict[str, Any]]:
    def cell(i: int) -> str:
        return _cell_str(row[i]) if i < len(row) else ""

    gmb_url = cell(0)
    name = cell(1)
    if not gmb_url or not name:
        return None

    rest = [cell(i) for i in range(2, len(row))]
    rating, reviews, category = _extract_rating_reviews_category(rest, name)

    lat, lon = _extract_lat_lon(gmb_url)

    address = phone = website = ""
    for s in rest:
        if not s or s.startswith("#"):
            continue
        if not address and _looks_like_address(s):
            address = s
        elif not phone and _is_phone(s):
            phone = _clean_phone(s)
        elif not website and _is_website(s):
            website = s

    return {
        "gmb_url": gmb_url, "name": name, "rating": rating, "reviews": reviews,
        "category": category, "address": address, "phone": phone,
        "website": website, "lat": lat, "lon": lon,
    }


def clean_gmb_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse a raw Google Maps scrape export and return deduped, clean business rows + stats."""
    if filename.lower().endswith(".csv"):
        rows = _read_csv_rows(file_bytes)
    else:
        rows = _read_xlsx_rows(file_bytes)

    if not rows:
        return {"businesses": [], "stats": {
            "total_rows": 0, "dropped_blank_or_noname": 0,
            "duplicates_removed": 0, "clean_rows": 0, "phone_recovered": 0,
        }}

    data_rows = rows[1:]  # first row is the header
    total_rows = len(data_rows)

    businesses = []
    seen_urls = set()
    dropped_blank = 0
    duplicates_removed = 0
    phone_recovered = 0

    for row in data_rows:
        row = list(row)
        if not any(str(c).strip() for c in row if c is not None):
            dropped_blank += 1
            continue

        biz = _parse_row(row)
        if biz is None:
            dropped_blank += 1
            continue

        if biz["gmb_url"] in seen_urls:
            duplicates_removed += 1
            continue
        seen_urls.add(biz["gmb_url"])

        if biz["phone"]:
            phone_recovered += 1
        businesses.append(biz)

    stats = {
        "total_rows": total_rows,
        "dropped_blank_or_noname": dropped_blank,
        "duplicates_removed": duplicates_removed,
        "clean_rows": len(businesses),
        "phone_recovered": phone_recovered,
    }
    return {"businesses": businesses, "stats": stats}


EXPORT_HEADERS = ["GMB URL", "Name", "Rating", "Reviews", "Category", "Address", "Phone", "Website", "Lat", "Lon"]


def build_export_excel(businesses: List[Dict[str, Any]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Clean Data"
    ws.append(EXPORT_HEADERS)
    for b in businesses:
        ws.append([
            b["gmb_url"], b["name"], b["rating"], b["reviews"], b["category"],
            b["address"], b["phone"], b["website"], b["lat"], b["lon"],
        ])
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
