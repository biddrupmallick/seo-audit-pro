"""
Nearby Finder — upload a business list (with lat/lon columns) and look up
the top N closest businesses to a given point or named business.
"""
import csv
import io
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl import Workbook

from analyzers.geo_match import haversine_miles

COLUMN_ALIASES = {
    "gmb_url": ["gmb_url", "gmb url", "google_maps_url", "maps_url", "google maps url", "url", "maps url"],
    "name": ["name", "business_name", "business name", "company", "company name"],
    "details": ["details", "detials", "description", "detail", "about"],
    "rating": ["rating", "star_rating", "stars", "avg_rating", "average rating"],
    "reviews": ["reviews", "review_count", "num_reviews", "number of reviews", "total reviews"],
    "category": ["category", "niche", "business_type", "business type", "type"],
    "address": ["address", "full_address", "street address"],
    "phone": ["phone", "phone_number", "telephone", "tel"],
    "website": ["website", "web", "site", "website_url"],
    "lat": ["lat", "latitude"],
    "lon": ["lan", "lon", "long", "longitude"],
    "owner_name": ["owner_name", "owner name", "owner", "contact", "contact name"],
    "email": ["email", "owner_email", "owner email", "e-mail"],
}


def _normalise_header(header: str) -> str:
    return header.strip().lower().replace("-", "_").replace(" ", "_")


def _map_columns(headers: List[str]) -> Dict[str, int]:
    normalised = [_normalise_header(h) for h in headers]
    mapping = {}
    for canon, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            alias_n = alias.replace(" ", "_")
            if alias_n in normalised:
                mapping[canon] = normalised.index(alias_n)
                break
    return mapping


def _extra_headers(headers: List[Any], known_indices: set) -> List[str]:
    """Original header text for columns not mapped to a known field, in file order."""
    result = []
    for i, h in enumerate(headers):
        if i in known_indices:
            continue
        h_clean = str(h).strip() if h is not None else ""
        if h_clean:
            result.append(h_clean)
    return result


def _row_to_dict(row: List[Any], col_map: Dict[str, int], headers: List[Any], known_indices: set) -> Dict[str, Any]:
    def get(key):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return ""
        val = row[idx]
        return str(val).strip() if val is not None else ""

    def get_float(key):
        raw = get(key)
        try:
            return float(raw)
        except (TypeError, ValueError):
            return None

    rating = get_float("rating")
    reviews = get_float("reviews")
    reviews = int(reviews) if reviews is not None else None

    extra = {}
    for i, h in enumerate(headers):
        if i in known_indices:
            continue
        h_clean = str(h).strip() if h is not None else ""
        if not h_clean:
            continue
        val = row[i] if i < len(row) else None
        extra[h_clean] = str(val).strip() if val is not None else ""

    return {
        "gmb_url": get("gmb_url"),
        "name": get("name"),
        "details": get("details"),
        "rating": rating,
        "reviews": reviews,
        "category": get("category"),
        "address": get("address"),
        "phone": get("phone"),
        "website": get("website"),
        "lat": get_float("lat"),
        "lon": get_float("lon"),
        "owner_name": get("owner_name"),
        "email": get("email"),
        "extra": extra,
    }


def parse_nearby_file(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    """Parse uploaded Excel/CSV into business dicts with lat/lon. Skips rows without valid coords.

    Any column not recognized as a known field (e.g. Facebook, LinkedIn, Instagram, X, YouTube)
    is kept under each business's "extra" dict and reported in "extra_headers", so new list
    formats work without code changes.
    """
    if filename.lower().endswith(".csv"):
        rows = _read_csv_rows(file_bytes)
    else:
        rows = _read_xlsx_rows(file_bytes)

    if not rows:
        return {"businesses": [], "extra_headers": []}

    headers = rows[0]
    col_map = _map_columns(headers)
    known_indices = set(col_map.values())
    extra_headers = _extra_headers(headers, known_indices)

    businesses = []
    for row in rows[1:]:
        biz = _row_to_dict(list(row), col_map, headers, known_indices)
        if not biz["name"]:
            continue
        if biz["lat"] is None or biz["lon"] is None:
            continue
        businesses.append(biz)
    return {"businesses": businesses, "extra_headers": extra_headers}


def _read_xlsx_rows(file_bytes: bytes) -> List[List[Any]]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _read_csv_rows(file_bytes: bytes) -> List[List[Any]]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def find_top_nearest(
    businesses: List[Dict[str, Any]],
    lat: float,
    lon: float,
    n: int = 5,
    exclude_index: Optional[int] = None,
) -> List[Dict[str, Any]]:
    """Return the n closest businesses to (lat, lon), each annotated with distance_miles."""
    results = []
    for i, biz in enumerate(businesses):
        if i == exclude_index:
            continue
        dist = haversine_miles(lat, lon, biz["lat"], biz["lon"])
        results.append({**biz, "distance_miles": round(dist, 2)})

    results.sort(key=lambda b: b["distance_miles"])
    return results[:n]


EXPORT_HEADERS = [
    "Rank", "Business Name", "Distance (mi)", "Rating", "Reviews", "Category",
    "Address", "Phone", "Website", "Owner Name", "Email", "Details",
    "Latitude", "Longitude", "GMB URL",
]


def _export_row(entry: Dict[str, Any], rank: Any, distance: Any, extra_headers: List[str]) -> List[Any]:
    base = [
        rank, entry.get("name", ""), distance, entry.get("rating"),
        entry.get("reviews"), entry.get("category", ""), entry.get("address", ""),
        entry.get("phone", ""), entry.get("website", ""), entry.get("owner_name", ""),
        entry.get("email", ""), entry.get("details", ""), entry.get("lat"), entry.get("lon"),
        entry.get("gmb_url", ""),
    ]
    extra = entry.get("extra") or {}
    return base + [extra.get(h, "") for h in extra_headers]


def build_export_excel(
    results: List[Dict[str, Any]],
    target: Optional[Dict[str, Any]] = None,
    extra_headers: Optional[List[str]] = None,
) -> bytes:
    """Build a downloadable Excel of nearest-business results. Target (if any) is pinned as the first row."""
    extra_headers = extra_headers or []
    wb = Workbook()
    ws = wb.active
    ws.title = "Nearby Results"

    ws.append(EXPORT_HEADERS + extra_headers)

    if target:
        ws.append(_export_row(target, "Search Center", 0, extra_headers))

    for i, r in enumerate(results, 1):
        ws.append(_export_row(r, i, r.get("distance_miles"), extra_headers))

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def find_business_by_name(businesses: List[Dict[str, Any]], name: str) -> Optional[int]:
    """Find index of business matching name (case-insensitive exact match, then substring)."""
    needle = name.strip().lower()
    if not needle:
        return None
    for i, biz in enumerate(businesses):
        if biz["name"].strip().lower() == needle:
            return i
    for i, biz in enumerate(businesses):
        if needle in biz["name"].strip().lower():
            return i
    return None
