import asyncio
import json
import os
import time
import traceback
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Dict, Any, Optional, List

import httpx
from fastapi import FastAPI, WebSocket, WebSocketDisconnect, HTTPException, BackgroundTasks, UploadFile, File, Form
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.requests import Request
from pydantic import BaseModel, HttpUrl

from config import REPORTS_DIR
from crawler.spider import Spider
from analyzers.technical import analyze_technical
from analyzers.onpage import analyze_onpage
from analyzers.schema import analyze_schema
from analyzers.aeo import analyze_aeo
from analyzers.geo import analyze_geo
from analyzers.performance import analyze_performance
from analyzers.images import analyze_images
from analyzers.local_seo import analyze_local_seo, enhance_with_gbp
from analyzers.conversion import analyze_conversion
from analyzers.content import analyze_content
from analyzers.ai_recommendations import generate_ai_recommendations
from analyzers.revenue_impact import calculate_revenue_impact
from analyzers.wayback import analyze_wayback
from analyzers.competitor import analyze_competitor
from analyzers.cold_email import generate_cold_emails
from analyzers.keyword_opportunities import analyze_keyword_opportunities
from analyzers.gbp import analyze_gbp
from analyzers.lead_score import calculate_lead_score
from analyzers.history import save_audit, get_history, build_progress
from analyzers.excel_parser import parse_excel, enrich_owner_info, extract_owner_info_single
from analyzers.trust_signals import analyze_trust_signals
from analyzers.mobile_screenshot import capture_mobile_screenshots
from analyzers.competitor_gap import analyze_competitor_gap
from analyzers.meta_rewrite import generate_meta_rewrites
from analyzers.roadmap import generate_roadmap
from analyzers.geo_match import find_nearest_competitors
from analyzers.review_analyzer import analyze_reviews_batch
from analyzers.website_email import enrich_businesses_with_website_emails
from analyzers.niche_blog import generate_blog_posts
from analyzers.ultra_email import generate_ultra_emails
from analyzers.text_cleaner import clean_review_text
from analyzers.file_prep import process_file as prep_process_file
from report.branding import load_branding, save_branding
from scoring.scorer import calculate_scores
from report.generator import generate_report, get_report_path

# ====== APP SETUP ======
@asynccontextmanager
async def lifespan(app):
    os.makedirs(REPORTS_DIR, exist_ok=True)
    print("=" * 60)
    print("  SEO Audit Pro is running!")
    print("  Open: http://localhost:8000")
    print("=" * 60)
    yield

app = FastAPI(
    title="SEO Audit Pro",
    description="Unlimited local SEO crawler and audit tool",
    version="1.0.0",
    lifespan=lifespan,
)

BASE_DIR = Path(__file__).parent

app.mount("/static", StaticFiles(directory=str(BASE_DIR / "static")), name="static")
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))
templates.env.globals["enumerate"] = enumerate
templates.env.globals["len"] = len

# ====== IN-MEMORY JOB STORE ======
# Structure: {job_id: {"status": str, "data": dict, "progress": int, "message": str, "ws": WebSocket}}
jobs: Dict[str, Dict[str, Any]] = {}
# WebSocket connections per job
ws_connections: Dict[str, WebSocket] = {}

# Bulk job store
bulk_jobs: Dict[str, Dict[str, Any]] = {}

# Upload pipeline store
upload_jobs: Dict[str, Dict[str, Any]] = {}


# ====== MODELS ======
class AnalyzeRequest(BaseModel):
    url: str
    competitor_urls: Optional[List[str]] = None
    gbp_url: Optional[str] = None
    gbp_competitor_urls: Optional[List[str]] = None


class BulkItem(BaseModel):
    url: str
    competitor_urls: Optional[List[str]] = None


class BulkRequest(BaseModel):
    items: List[BulkItem]


# ====== ROUTES ======
@app.get("/", response_class=HTMLResponse)
async def index(request: Request):
    is_local = request.url.hostname in ('localhost', '127.0.0.1')
    return templates.TemplateResponse(request, "index.html", {"is_local": is_local})


@app.get("/settings", response_class=HTMLResponse)
async def settings_page(request: Request):
    return templates.TemplateResponse(request, "settings.html", {
        "branding": load_branding(),
    })


# ====== REVIEW CLEANER ======
@app.get("/review-cleaner", response_class=HTMLResponse)
async def review_cleaner_page(request: Request):
    return templates.TemplateResponse(request, "review_cleaner.html", {})


@app.post("/api/review-cleaner/clean")
async def review_cleaner_clean(request: Request):
    body = await request.json()
    raw = body.get("text", "")
    cleaned, stats = clean_review_text(raw)
    return {"cleaned": cleaned, "stats": stats}


@app.post("/api/review-cleaner/analyse")
async def review_cleaner_analyse(request: Request):
    from analyzers.ollama_client import ask
    body = await request.json()
    text = body.get("text", "").strip()
    if not text:
        return {"analysis": {}}

    prompt = f"""You are a market research analyst specialising in local business reviews.

REVIEW TEXT:
{text[:6000]}

Analyse these reviews and respond in EXACTLY this format:

TOP_PRAISE: [3 most common things customers praise, comma-separated]
TOP_COMPLAINTS: [3 most common complaints, comma-separated]
STAFF_PATTERNS: [behaviours or traits mentioned about staff, comma-separated]
SERVICE_KEYWORDS: [top 5 services customers actually mention, comma-separated]
DIFFERENTIATORS: [what separates positive from negative experiences, one sentence]
CUSTOMER_LANGUAGE: [3-5 exact words/phrases customers use repeatedly, in quotes]
KEY_INSIGHT: [single most actionable finding from these reviews, one sentence]"""

    raw = await asyncio.to_thread(ask, prompt, 500, 0.4)
    parsed = {}
    for line in raw.splitlines():
        if ":" in line:
            k, _, v = line.partition(":")
            parsed[k.strip()] = v.strip()
    return {"analysis": parsed}


# ====== FILE PREP ======
file_prep_jobs: Dict[str, Dict[str, Any]] = {}


@app.get("/file-prep", response_class=HTMLResponse)
async def file_prep_page(request: Request):
    return templates.TemplateResponse(request, "file_prep.html", {})


@app.post("/api/file-prep/preview")
async def file_prep_preview(file: UploadFile = File(...)):
    import openpyxl, io
    contents = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
    ws = wb.active
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, min(7, ws.max_row + 1)):
        row = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if any(v for v in row):
            rows.append([str(v)[:60] if v else "" for v in row])
    return {"headers": [str(h) if h else "" for h in headers], "rows": rows}


@app.post("/api/file-prep/start")
async def file_prep_start(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    gmb_col: int = Form(2),
    name_col: int = Form(3),
):
    job_id = str(uuid.uuid4())
    contents = await file.read()
    file_prep_jobs[job_id] = {
        "status": "running",
        "current": 0,
        "total": 0,
        "message": "Starting…",
        "result": None,
    }

    def run(jid, data, gc, nc):
        def cb(current, total, msg):
            file_prep_jobs[jid]["current"] = current
            file_prep_jobs[jid]["total"] = total
            file_prep_jobs[jid]["message"] = msg
        try:
            result = prep_process_file(data, gc, nc, progress_callback=cb)
            file_prep_jobs[jid]["status"] = "complete"
            file_prep_jobs[jid]["result"] = result
        except Exception as e:
            file_prep_jobs[jid]["status"] = "error"
            file_prep_jobs[jid]["message"] = str(e)

    background_tasks.add_task(run, job_id, contents, gmb_col, name_col)
    return {"job_id": job_id}


@app.get("/api/file-prep/status/{job_id}")
async def file_prep_status(job_id: str):
    if job_id not in file_prep_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = file_prep_jobs[job_id]
    return {
        "status": job["status"],
        "current": job["current"],
        "total": job["total"],
        "message": job["message"],
    }


@app.get("/api/file-prep/download/{job_id}")
async def file_prep_download(job_id: str):
    if job_id not in file_prep_jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = file_prep_jobs[job_id]
    if job["status"] != "complete" or not job["result"]:
        raise HTTPException(status_code=400, detail="File not ready")
    return Response(
        content=job["result"],
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=clean_data.xlsx"},
    )


@app.post("/settings")
async def save_settings(request: Request):
    form = await request.form()
    data = {k: v for k, v in form.items()}
    save_branding(data)
    return templates.TemplateResponse(request, "settings.html", {
        "branding": load_branding(),
        "saved": True,
    })


@app.post("/analyze")
async def start_analysis(request: AnalyzeRequest, background_tasks: BackgroundTasks):
    """Start a new SEO analysis job."""
    url = request.url.strip()

    # Basic URL validation
    if not url.startswith(("http://", "https://")):
        url = "https://" + url

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status": "pending",
        "url": url,
        "created_at": datetime.now().isoformat(),
        "progress": 0,
        "message": "Queued",
        "data": None,
        "error": None,
    }

    raw_competitors = request.competitor_urls or []
    competitor_urls = []
    for cu in raw_competitors[:5]:
        cu = cu.strip()
        if cu:
            if not cu.startswith(("http://", "https://")):
                cu = "https://" + cu
            competitor_urls.append(cu)

    gbp_url = (request.gbp_url or "").strip() or None
    gbp_competitor_urls = [u.strip() for u in (request.gbp_competitor_urls or []) if u.strip()][:5]

    background_tasks.add_task(run_analysis, job_id, url, competitor_urls, gbp_url, gbp_competitor_urls)

    return {"job_id": job_id, "status": "started"}


@app.websocket("/ws/{job_id}")
async def websocket_endpoint(websocket: WebSocket, job_id: str):
    """WebSocket endpoint for live progress updates."""
    await websocket.accept()
    ws_connections[job_id] = websocket

    try:
        # If job already done, send result immediately
        if job_id in jobs:
            job = jobs[job_id]
            if job["status"] == "complete":
                await websocket.send_json({
                    "type": "complete",
                    "data": job["data"],
                })
                return
            elif job["status"] == "error":
                await websocket.send_json({
                    "type": "error",
                    "message": job.get("error", "Unknown error"),
                })
                return

        # Keep connection alive while job runs
        while True:
            try:
                # Check for messages from client (keepalive)
                await asyncio.wait_for(websocket.receive_text(), timeout=1.0)
            except asyncio.TimeoutError:
                pass
            except WebSocketDisconnect:
                break

            # Check job status
            if job_id not in jobs:
                await websocket.send_json({"type": "error", "message": "Job not found"})
                break

            job = jobs[job_id]
            if job["status"] == "complete":
                await websocket.send_json({
                    "type": "complete",
                    "data": job["data"],
                })
                break
            elif job["status"] == "error":
                await websocket.send_json({
                    "type": "error",
                    "message": job.get("error", "Unknown error"),
                })
                break

    except WebSocketDisconnect:
        pass
    except Exception as e:
        try:
            await websocket.send_json({"type": "error", "message": str(e)})
        except Exception:
            pass
    finally:
        ws_connections.pop(job_id, None)


@app.get("/report/{job_id}")
async def download_report(job_id: str):
    """Download the generated report (PDF or HTML)."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs[job_id]
    if job["status"] != "complete":
        raise HTTPException(status_code=400, detail="Analysis not complete yet")

    report_path = get_report_path(job_id)
    if not report_path:
        raise HTTPException(status_code=404, detail="Report file not found")

    if report_path.endswith(".pdf"):
        return FileResponse(
            report_path,
            media_type="application/pdf",
            filename=f"seo-audit-{job_id[:8]}.pdf",
        )
    else:
        return FileResponse(
            report_path,
            media_type="text/html",
            filename=f"seo-audit-{job_id[:8]}.html",
        )


@app.get("/status/{job_id}")
async def get_job_status(job_id: str):
    """Get the current status of a job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    job = jobs[job_id].copy()
    # Don't return full data in status check
    job.pop("data", None)
    return job


# ====== BULK AUDIT ROUTES ======
@app.get("/bulk", response_class=HTMLResponse)
async def bulk_page(request: Request):
    is_local = request.url.hostname in ('localhost', '127.0.0.1')
    return templates.TemplateResponse(request, "bulk.html", {"is_local": is_local})


@app.post("/bulk/start")
async def start_bulk(request: BulkRequest, background_tasks: BackgroundTasks):
    bulk_id = str(uuid.uuid4())
    items = []
    for item in request.items:
        url = item.url.strip()
        if not url:
            continue
        if not url.startswith(("http://", "https://")):
            url = "https://" + url
        comp_urls = []
        for cu in (item.competitor_urls or [])[:3]:
            cu = cu.strip()
            if cu:
                if not cu.startswith(("http://", "https://")):
                    cu = "https://" + cu
                comp_urls.append(cu)
        job_id = str(uuid.uuid4())
        items.append({
            "url": url,
            "competitor_urls": comp_urls,
            "job_id": job_id,
            "status": "pending",
            "lead_score": None,
            "report_path": None,
            "error": None,
            "progress": 0,
            "message": "Waiting…",
        })

    bulk_jobs[bulk_id] = {
        "bulk_id": bulk_id,
        "status": "running",
        "created_at": datetime.now().isoformat(),
        "items": items,
        "current_index": 0,
    }
    background_tasks.add_task(run_bulk, bulk_id)
    return {"bulk_id": bulk_id}


@app.get("/bulk/status/{bulk_id}")
async def bulk_status(bulk_id: str):
    if bulk_id not in bulk_jobs:
        raise HTTPException(status_code=404, detail="Bulk job not found")
    bj = bulk_jobs[bulk_id]
    # Return status without heavy data
    items_summary = []
    for it in bj["items"]:
        pct = it.get("progress", 0)
        eta_seconds = None
        started_at = it.get("started_at")
        if started_at and pct > 2 and it["status"] == "running":
            elapsed = time.time() - started_at
            eta_seconds = int(elapsed / pct * (100 - pct))
        items_summary.append({
            "url": it["url"],
            "job_id": it["job_id"],
            "status": it["status"],
            "progress": pct,
            "message": it.get("message", ""),
            "eta_seconds": eta_seconds,
            "error": it.get("error"),
            "lead_score": it.get("lead_score"),
            "report_path": it.get("report_path"),
        })
    return {
        "bulk_id": bulk_id,
        "status": bj["status"],
        "current_index": bj.get("current_index", 0),
        "total": len(bj["items"]),
        "items": items_summary,
    }


async def run_bulk(bulk_id: str):
    """Run audits sequentially for all items in a bulk job."""
    bj = bulk_jobs[bulk_id]
    items = bj["items"]

    for idx, item in enumerate(items):
        bj["current_index"] = idx
        item["status"] = "running"
        item["progress"] = 2
        item["message"] = "Starting…"
        item["started_at"] = time.time()

        job_id = item["job_id"]
        # Register in the main jobs store so run_analysis works
        jobs[job_id] = {
            "status": "pending",
            "url": item["url"],
            "created_at": datetime.now().isoformat(),
            "progress": 0,
            "message": "Queued",
            "data": None,
            "error": None,
        }

        # Mirror progress updates into bulk item
        original_send = None

        async def bulk_progress_hook(jid, pct, msg, _item=item):
            _item["progress"] = pct
            _item["message"] = msg

        # Temporarily patch send_progress for this job
        async def _patched_send(jid, pct, msg):
            jobs[jid]["progress"] = pct
            jobs[jid]["message"] = msg
            item["progress"] = pct
            item["message"] = msg

        try:
            await run_analysis(job_id, item["url"], item["competitor_urls"])
            job = jobs[job_id]
            if job["status"] == "complete" and job.get("data"):
                item["status"] = "complete"
                item["lead_score"] = job["data"].get("lead_score", {})
                item["report_path"] = job["data"].get("report_path", "")
                item["progress"] = 100
                item["message"] = "Complete"
            else:
                item["status"] = "error"
                item["error"] = job.get("error", "Unknown error")
        except Exception as e:
            item["status"] = "error"
            item["error"] = str(e)

    bj["status"] = "complete"
    bj["current_index"] = len(items)


# ====== UPLOAD PIPELINE ROUTES ======
@app.get("/upload", response_class=HTMLResponse)
async def upload_page(request: Request):
    is_local = request.url.hostname in ('localhost', '127.0.0.1')
    return templates.TemplateResponse(request, "upload.html", {"is_local": is_local})


@app.post("/upload/start")
async def start_upload(background_tasks: BackgroundTasks, file: UploadFile = File(...)):
    content = await file.read()
    try:
        businesses = parse_excel(content, file.filename)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not parse file: {e}")

    if not businesses:
        raise HTTPException(status_code=400, detail="No businesses found in file. Check column names.")

    upload_id = str(uuid.uuid4())
    upload_jobs[upload_id] = {
        "upload_id": upload_id,
        "status": "running",
        "step": "Starting…",
        "step_index": 0,
        "total_steps": 6,
        "businesses": businesses,
        "total": len(businesses),
        "created_at": datetime.now().isoformat(),
        "audit_progress": {"current": 0, "total": 0, "current_name": ""},
        "result": None,
        "error": None,
    }
    background_tasks.add_task(run_upload_pipeline, upload_id)
    return {"upload_id": upload_id, "total": len(businesses)}


@app.get("/upload/status/{upload_id}")
async def upload_status(upload_id: str):
    if upload_id not in upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    job = upload_jobs[upload_id]
    return {
        "upload_id": upload_id,
        "status": job["status"],
        "step": job["step"],
        "step_index": job["step_index"],
        "total_steps": job["total_steps"],
        "total": job["total"],
        "error": job.get("error"),
        "has_result": job.get("result") is not None,
        "audit_progress": job.get("audit_progress", {}),
        "partial_results": _make_serializable(job.get("partial_results", [])),
    }


@app.get("/upload/results/{upload_id}")
async def upload_results(upload_id: str):
    if upload_id not in upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    job = upload_jobs[upload_id]
    if job["status"] != "complete":
        raise HTTPException(status_code=400, detail="Pipeline not complete yet")
    return JSONResponse(content=_make_serializable(job["result"]))


@app.get("/upload/export-excel/{upload_id}")
async def export_excel(upload_id: str):
    """Export all cold emails + contact info to a formatted Excel file."""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse

    if upload_id not in upload_jobs:
        raise HTTPException(status_code=404, detail="Upload job not found")
    job = upload_jobs[upload_id]
    if job["status"] != "complete":
        raise HTTPException(status_code=400, detail="Pipeline not complete yet")

    result = job["result"]
    emails = result.get("emails", [])
    audit_results = {r["website"]: r for r in result.get("audit_results", [])}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cold Emails"

    # Styles
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    alt_fill = PatternFill("solid", fgColor="F0F4FF")
    border = Border(
        bottom=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
    )
    wrap = Alignment(wrap_text=True, vertical="top")

    headers = [
        "#", "Business Name", "Owner Name", "Contact Email", "Email Source",
        "Website", "Subject Line", "Email Body (2 sentences)",
        "Nearest Competitor", "Distance (mi)", "Lead Score", "Lead Tier",
    ]
    col_widths = [4, 28, 20, 32, 12, 30, 40, 70, 28, 13, 11, 14]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for i, e in enumerate(emails, 1):
        row = i + 1
        fill = alt_fill if i % 2 == 0 else None
        audit = audit_results.get(e.get("website", ""), {})
        ls = audit.get("lead_score", {})

        src = "owner" if e.get("owner_email") else ("website" if e.get("website_email") else "")
        values = [
            i,
            e.get("name", ""),
            e.get("owner_name", ""),
            e.get("contact_email", ""),
            src,
            e.get("website", ""),
            e.get("subject", ""),
            e.get("body", ""),
            e.get("nearest_competitor", ""),
            e.get("distance", ""),
            ls.get("score", ""),
            ls.get("tier", ""),
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.alignment = wrap
            cell.border = border
            if fill:
                cell.fill = fill

        ws.row_dimensions[row].height = 55

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=cold_emails_{upload_id[:8]}.xlsx"},
    )


def _estimate_dist(rating: float, count: int) -> dict:
    if not rating or not count:
        return {}
    five = max(0, int(count * min(0.95, (rating - 1) / 4 * 0.9 + 0.1)))
    one = max(0, int(count * max(0.01, (5 - rating) / 4 * 0.15)))
    remaining = max(0, count - five - one)
    four = int(remaining * 0.5)
    three = int(remaining * 0.3)
    two = remaining - four - three
    return {"5": five, "4": four, "3": three, "2": two, "1": one}


def _quick_review_insight(name: str, reviews_text: str) -> Dict[str, str]:
    """Run a fast Ollama call to extract praise/complaint from a competitor's reviews."""
    if not reviews_text or len(reviews_text.strip()) < 30:
        return {"praise": "", "complaint": ""}
    try:
        from analyzers.ollama_client import ask
        prompt = (
            f"Analyse these customer reviews for '{name}'.\n\n"
            f"REVIEWS:\n{reviews_text[:2000]}\n\n"
            "Output EXACTLY 2 lines:\n"
            "GOOD_AT: [one short phrase — what customers consistently praise]\n"
            "BAD_AT: [one short phrase — what customers consistently complain about, or 'nothing notable']"
        )
        raw = ask(prompt, max_tokens=60, temperature=0.3)
        result = {"praise": "", "complaint": ""}
        for line in raw.splitlines():
            line = line.strip()
            if line.startswith("GOOD_AT:"):
                result["praise"] = line[8:].strip()
            elif line.startswith("BAD_AT:"):
                result["complaint"] = line[7:].strip()
        return result
    except Exception:
        return {"praise": "", "complaint": ""}


def _gbp_from_excel(biz: Dict) -> Dict:
    """Build GBP-compatible dict from Excel row data."""
    competitors = biz.get("nearest_competitors") or []
    reviews = biz.get("reviews") or 0
    rating = biz.get("rating")
    group_analysis = biz.get("_group_analysis") or {}

    score = 0
    if rating:
        score += int((rating / 5.0) * 40)
    if reviews >= 100: score += 30
    elif reviews >= 50: score += 22
    elif reviews >= 20: score += 15
    elif reviews >= 5: score += 8
    if biz.get("phone"): score += 10
    if biz.get("address"): score += 10
    if biz.get("website"): score += 10
    score = min(100, score)

    comp_counts = [c.get("reviews") or 0 for c in competitors]
    top_comp = max(comp_counts) if comp_counts else 0
    avg_comp = round(sum(comp_counts) / max(len(comp_counts), 1)) if comp_counts else 0
    review_gap = max(0, top_comp - reviews)

    # Per-business review insight from the business's own reviews_text
    own_insight = _quick_review_insight(biz.get("name", ""), biz.get("reviews_text", "") or "")
    # Fall back to group-level analysis if per-business reviews_text is empty
    insights = {
        "PRAISE_THEMES": own_insight.get("praise", "") or group_analysis.get("TOP_PRAISE", ""),
        "COMPLAINT_THEMES": own_insight.get("complaint", "") or group_analysis.get("TOP_COMPLAINTS", ""),
        "TOP_ACTION": group_analysis.get("KEY_INSIGHT", ""),
        "REVIEW_ASK": f"Would you mind leaving us a quick Google review? It helps other {biz.get('category', 'customers')} find us.",
        "RESPONSE_SCRIPT": "Thank you for your feedback. We take all reviews seriously and will use this to improve our service.",
    }

    # Build comparison table from nearest competitors
    comparison = {}
    if competitors:
        comp_names = [c.get("name", f"Competitor {i+1}")[:25] for i, c in enumerate(competitors)]
        rows = [
            {"label": "Rating", "client": str(rating) + "★" if rating else "—",
             "competitors": [str(c.get("rating", "—")) + "★" if c.get("rating") else "—" for c in competitors]},
            {"label": "Reviews", "client": str(reviews),
             "competitors": [str(c.get("reviews") or "—") for c in competitors]},
            {"label": "Distance", "client": "📍 Client",
             "competitors": [str(c.get("distance_miles", "?")) + " mi" for c in competitors]},
            {"label": "Website", "client": "✓" if biz.get("website") else "✗",
             "competitors": ["✓" if c.get("website") else "✗" for c in competitors]},
        ]
        comparison = {"competitor_names": comp_names, "rows": rows}

    return {
        "available": True,
        "source": "excel",
        "client": {
            "name": biz.get("name", ""),
            "rating": rating,
            "review_count": reviews,
            "category": biz.get("category", ""),
            "address": biz.get("address", ""),
            "phone": biz.get("phone", ""),
            "website": biz.get("website", ""),
            "score": score,
            "has_hours": True,
            "has_photos": True,
            "has_description": True,
            "has_posts": False,
            "photo_count": None,
            "issues": [],
            "wins": [],
        },
        "competitors": [
            {
                "name": c.get("name", ""),
                "rating": c.get("rating"),
                "review_count": c.get("reviews"),
                "reviews": c.get("reviews"),
                "distance_miles": c.get("distance_miles"),
                "website": c.get("website", ""),
                **_quick_review_insight(c.get("name", ""), c.get("reviews_text", "")),
            }
            for c in competitors
        ],
        "comparison": comparison,
        "review_intel": {
            "distribution": _estimate_dist(rating, reviews) if rating and reviews else {},
            "count": reviews,
            "review_gap": review_gap,
            "top_competitor_reviews": top_comp,
            "competitor_avg_count": avg_comp,
            "months_to_close_gap": round(review_gap / 4) if review_gap > 0 else 0,
            "insights": insights,
        },
    }


async def run_upload_pipeline(upload_id: str):
    """Full pipeline: prep → per-business (email+PDF) → blog → finalise."""
    job = upload_jobs[upload_id]
    businesses = job["businesses"]

    # Cache: survive Colab disconnects — already-completed businesses are skipped on restart
    cache_path = os.path.join("data", f"{upload_id}_cache.json")
    os.makedirs("data", exist_ok=True)
    try:
        with open(cache_path) as f:
            cache = json.load(f)
    except Exception:
        cache = {}

    def _save_cache():
        try:
            with open(cache_path, "w") as f:
                json.dump(cache, f)
        except Exception:
            pass

    try:
        # Step 1: Owner info extraction + geo-match (fast, no AI)
        job["step_index"] = 1
        needs_enrichment = [b for b in businesses if b.get("owner_info")]
        for ei, biz in enumerate(needs_enrichment):
            biz_name = biz.get("name") or biz.get("website") or f"Business {ei+1}"
            job["step"] = f"Reading owner info for {biz_name}… ({ei+1}/{len(needs_enrichment)})"
            extracted = await asyncio.to_thread(extract_owner_info_single, biz)
            if not biz.get("owner_name"):
                biz["owner_name"] = extracted["owner_name"]
            if not biz.get("owner_email"):
                biz["owner_email"] = extracted["owner_email"]

        job["step"] = "Extracting coordinates & matching nearest competitors…"
        enriched = await asyncio.to_thread(find_nearest_competitors, businesses)

        # Step 2: Per-business loop — email + PDF one at a time
        job["step_index"] = 2
        job["partial_results"] = []
        job["audit_progress"] = {"current": 0, "total": len(enriched), "current_name": ""}
        audit_results = []
        all_emails = []
        niche_groups = {}  # accumulate data for blog posts

        from analyzers.ultra_email import _generate_email_for_business
        from analyzers.website_email import get_best_contact_email

        for i, biz in enumerate(enriched):
            biz_name = biz.get("name", f"Business {i+1}")
            cache_key = biz_name or str(i)

            job["step"] = f"Processing {biz_name} ({i+1}/{len(enriched)})…"
            job["audit_progress"] = {"current": i + 1, "total": len(enriched), "current_name": biz_name}

            # Restore from cache if this business was already completed
            if cache_key in cache:
                row = cache[cache_key]
                if row.get("job_id") and row.get("report_path"):
                    jobs[row["job_id"]] = {
                        "status": "complete", "url": row.get("website", ""),
                        "data": {"report_path": row["report_path"]}, "error": None,
                    }
                audit_results.append(row)
                job["partial_results"].append(row)
                if row.get("subject"):
                    all_emails.append({k: row.get(k, "") for k in [
                        "name", "owner_name", "owner_email", "website_email",
                        "contact_email", "website", "subject", "body",
                        "nearest_competitor", "distance",
                    ]})
                continue

            # Scrape website email (fast HTTP, no AI)
            website = (biz.get("website") or "").strip()
            if website:
                if not website.startswith(("http://", "https://")):
                    website = "https://" + website
                try:
                    biz["website_email"] = await asyncio.to_thread(get_best_contact_email, website)
                except Exception:
                    biz["website_email"] = ""
            else:
                biz["website_email"] = ""

            owner_email = biz.get("owner_email", "")
            website_email = biz.get("website_email", "")
            comp = (biz.get("nearest_competitors") or [{}])[0]

            # Cold email (uses competitor data already available — no extra Ollama call)
            email_result = await asyncio.to_thread(_generate_email_for_business, biz, {})
            email_row = {
                "name": biz_name,
                "owner_name": biz.get("owner_name", ""),
                "owner_email": owner_email,
                "website_email": website_email,
                "contact_email": owner_email or website_email,
                "website": website,
                "subject": email_result["subject"],
                "body": email_result["body"],
                "nearest_competitor": comp.get("name", ""),
                "distance": comp.get("distance_miles", ""),
            }
            all_emails.append(email_row)

            # Accumulate niche data for blog posts
            niche_key = f"{(biz.get('category') or '').strip()} | {(biz.get('state') or '').strip()}"
            if niche_key not in niche_groups:
                niche_groups[niche_key] = {
                    "category": (biz.get("category") or "").strip(),
                    "state": (biz.get("state") or "").strip(),
                    "count": 0, "ratings": [], "reviews": [],
                }
            niche_groups[niche_key]["count"] += 1
            if biz.get("rating"):
                try: niche_groups[niche_key]["ratings"].append(float(biz["rating"]))
                except Exception: pass
            if biz.get("reviews"):
                try: niche_groups[niche_key]["reviews"].append(int(biz["reviews"]))
                except Exception: pass

            # PDF: full website audit OR no-website pitch
            audit_job_id = str(uuid.uuid4())

            if website:
                jobs[audit_job_id] = {
                    "status": "pending", "url": website,
                    "created_at": datetime.now().isoformat(),
                    "progress": 0, "message": "Queued", "data": None, "error": None,
                }
                comp_websites = []
                for c in biz.get("nearest_competitors") or []:
                    cw = (c.get("website") or "").strip()
                    if cw and cw != biz.get("website", "").strip():
                        if not cw.startswith(("http://", "https://")):
                            cw = "https://" + cw
                        comp_websites.append(cw)

                try:
                    await run_analysis(audit_job_id, website, comp_websites[:3], biz_data=biz)
                    aj = jobs[audit_job_id]
                    if aj["status"] == "complete" and aj.get("data"):
                        row = {
                            "business_name": biz_name,
                            "owner_name": biz.get("owner_name", ""),
                            "owner_email": owner_email,
                            "website_email": website_email,
                            "contact_email": owner_email or website_email,
                            "category": biz.get("category", ""),
                            "website": website,
                            "job_id": audit_job_id,
                            "lead_score": aj["data"].get("lead_score", {}),
                            "report_path": aj["data"].get("report_path", ""),
                            "subject": email_result["subject"],
                            "body": email_result["body"],
                        }
                    else:
                        row = {
                            "business_name": biz_name,
                            "owner_name": biz.get("owner_name", ""),
                            "owner_email": owner_email,
                            "website_email": website_email,
                            "contact_email": owner_email or website_email,
                            "website": website,
                            "job_id": audit_job_id,
                            "error": aj.get("error", "Unknown error"),
                            "subject": email_result["subject"],
                            "body": email_result["body"],
                        }
                except Exception as e:
                    row = {
                        "business_name": biz_name,
                        "website": website,
                        "job_id": audit_job_id,
                        "error": str(e),
                        "subject": email_result["subject"],
                        "body": email_result["body"],
                    }
            else:
                # No website — generate pitch PDF directly
                gbp_data = await asyncio.to_thread(_gbp_from_excel, biz)
                try:
                    report_path = await asyncio.to_thread(
                        generate_report,
                        job_id=audit_job_id,
                        root_url="",
                        total_pages=0,
                        scores={}, technical={}, onpage={}, schema={},
                        aeo={}, geo={}, performance={}, images={},
                        gbp=gbp_data,
                    )
                    jobs[audit_job_id] = {"status": "complete", "url": "", "data": {"report_path": report_path}, "error": None}
                    row = {
                        "business_name": biz_name,
                        "owner_name": biz.get("owner_name", ""),
                        "owner_email": owner_email,
                        "website_email": website_email,
                        "contact_email": owner_email or website_email,
                        "category": biz.get("category", ""),
                        "website": "",
                        "job_id": audit_job_id,
                        "lead_score": {},
                        "report_path": report_path,
                        "no_website": True,
                        "subject": email_result["subject"],
                        "body": email_result["body"],
                    }
                except Exception as e:
                    jobs[audit_job_id] = {"status": "error", "url": "", "data": None, "error": str(e)}
                    row = {
                        "business_name": biz_name,
                        "website": "",
                        "job_id": audit_job_id,
                        "error": str(e),
                        "no_website": True,
                        "subject": email_result["subject"],
                        "body": email_result["body"],
                    }

            cache[cache_key] = row
            _save_cache()
            audit_results.append(row)
            job["partial_results"].append(row)

        audit_results.sort(
            key=lambda x: (x.get("lead_score") or {}).get("score", 0),
            reverse=True,
        )

        # Step 3: Blog posts (per niche, uses data collected during loop)
        job["step_index"] = 3
        blog_posts = {}
        niche_list = list(niche_groups.items())
        for gi, (key, grp) in enumerate(niche_list):
            job["step"] = f"Writing blog post for {grp['category']} | {grp['state']} ({gi+1}/{len(niche_list)})…"
            avg_rating = round(sum(grp["ratings"]) / len(grp["ratings"]), 1) if grp["ratings"] else 0
            avg_reviews = int(sum(grp["reviews"]) / len(grp["reviews"])) if grp["reviews"] else 0
            if avg_rating or grp["count"] > 0:
                try:
                    posts = await asyncio.to_thread(
                        generate_blog_posts,
                        grp["category"], grp["state"],
                        grp["count"], avg_rating, avg_reviews, {},
                    )
                    blog_posts[key] = posts
                except Exception:
                    pass

        # Step 4: Finalise
        job["step_index"] = 4
        job["step"] = "Finalising results…"

        located = [b for b in enriched if b.get("latlon")]
        job["result"] = {
            "total": len(enriched),
            "located": len(located),
            "audited": len([r for r in audit_results if not r.get("error")]),
            "businesses": enriched,
            "review_analysis": {},
            "blog_posts": blog_posts,
            "emails": all_emails,
            "audit_results": audit_results,
        }
        job["status"] = "complete"
        job["step"] = "Complete!"

    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)
        job["step"] = f"Error: {e}"
        print(f"[Upload {upload_id}] Error: {traceback.format_exc()}")


# ====== BACKGROUND ANALYSIS TASK ======
async def send_progress(job_id: str, percent: int, message: str):
    """Send a progress update via WebSocket."""
    jobs[job_id]["progress"] = percent
    jobs[job_id]["message"] = message

    eta_seconds = None
    started_at = jobs[job_id].get("started_at")
    if started_at and percent > 2:
        elapsed = time.time() - started_at
        eta_seconds = int(elapsed / percent * (100 - percent))

    if job_id in ws_connections:
        try:
            await ws_connections[job_id].send_json({
                "type": "progress",
                "percent": percent,
                "message": message,
                "eta_seconds": eta_seconds,
            })
        except Exception:
            pass


async def run_analysis(job_id: str, url: str, competitor_urls: Optional[List[str]] = None, gbp_url: Optional[str] = None, gbp_competitor_urls: Optional[List[str]] = None, biz_data: Optional[Dict] = None):
    """Main background task: crawl + analyze + generate report."""
    try:
        jobs[job_id]["status"] = "running"
        jobs[job_id]["started_at"] = time.time()
        parsed_domain = url.replace("https://", "").replace("http://", "").split("/")[0]
        await send_progress(job_id, 2, f"Starting crawl of {url}…")

        # ====== CRAWL ======
        crawled_pages = []
        pages_found = [0]

        async def progress_cb(count, max_pages, current_url):
            pages_found[0] = count
            pct = min(45, int((count / max_pages) * 45))
            short_url = current_url[:60] + "…" if len(current_url) > 60 else current_url
            await send_progress(job_id, pct, f"Crawling page {count}: {short_url}")

        spider = Spider(url, progress_callback=progress_cb)
        crawled_pages = await spider.crawl()

        total_pages = len(crawled_pages)
        await send_progress(job_id, 47, f"Crawled {total_pages} pages. Running analyzers…")

        # ====== ANALYZE ======
        await send_progress(job_id, 50, "Analyzing technical SEO…")
        technical = await asyncio.to_thread(analyze_technical, crawled_pages)

        await send_progress(job_id, 57, "Analyzing on-page SEO…")
        onpage = await asyncio.to_thread(analyze_onpage, crawled_pages)

        await send_progress(job_id, 63, "Analyzing schema markup…")
        schema = await asyncio.to_thread(analyze_schema, crawled_pages)

        await send_progress(job_id, 69, "Analyzing AEO readiness…")
        aeo = await asyncio.to_thread(analyze_aeo, crawled_pages)

        await send_progress(job_id, 74, "Analyzing GEO readiness…")
        geo = await asyncio.to_thread(analyze_geo, crawled_pages)

        await send_progress(job_id, 79, "Analyzing performance…")
        performance = await asyncio.to_thread(analyze_performance, crawled_pages)

        await send_progress(job_id, 83, "Analyzing images…")
        images = await asyncio.to_thread(analyze_images, crawled_pages)

        await send_progress(job_id, 85, "Analyzing local SEO…")
        local_seo = await asyncio.to_thread(analyze_local_seo, crawled_pages)
        if biz_data:
            local_seo = enhance_with_gbp(local_seo, biz_data)

        await send_progress(job_id, 87, "Analyzing conversion optimization…")
        conversion = await asyncio.to_thread(analyze_conversion, crawled_pages)

        await send_progress(job_id, 89, "Analyzing content quality…")
        content = await asyncio.to_thread(analyze_content, crawled_pages)

        await send_progress(job_id, 90, "Finding keyword opportunities…")
        keywords = await asyncio.to_thread(analyze_keyword_opportunities, crawled_pages)

        await send_progress(job_id, 91, "Auditing trust signals…")
        trust_signals = await asyncio.to_thread(analyze_trust_signals, crawled_pages, parsed_domain)

        await send_progress(job_id, 91, "Rewriting weak meta descriptions…")
        meta_rewrites = await asyncio.to_thread(generate_meta_rewrites, crawled_pages, parsed_domain)

        await send_progress(job_id, 91, "Capturing mobile screenshots…")
        comp_url_for_screenshot = competitor_urls[0] if competitor_urls else None
        mobile_screenshots = await asyncio.to_thread(capture_mobile_screenshots, url, comp_url_for_screenshot, job_id)

        await send_progress(job_id, 91, "Calculating scores…")
        scores = calculate_scores(technical, onpage, schema, aeo, geo, performance, images, local_seo, conversion, content, total_pages)

        # Competitor analysis (optional — run all concurrently)
        competitors = []
        if competitor_urls:
            await send_progress(job_id, 91, f"Crawling {len(competitor_urls)} competitor(s)… (this may take a minute)")
            tasks = [
                analyze_competitor(
                    cu, parsed_domain,
                    scores, technical, onpage, schema, aeo, geo,
                    performance, images, local_seo, conversion, content,
                )
                for cu in competitor_urls
            ]
            results = await asyncio.gather(*tasks, return_exceptions=True)
            competitors = [r for r in results if isinstance(r, dict) and r.get("available")]

        # Save audit to history DB
        await asyncio.to_thread(save_audit, parsed_domain, scores)
        audit_history = await asyncio.to_thread(get_history, parsed_domain)
        progress_data = build_progress(audit_history)

        await send_progress(job_id, 91, "Checking Wayback Machine history…")
        wayback = await analyze_wayback(parsed_domain)

        # GBP: from Excel data (no scraping) or from provided URL
        gbp = {}
        if biz_data:
            gbp = _gbp_from_excel(biz_data)
        elif gbp_url:
            await send_progress(job_id, 91, "Auditing Google Business Profile…")
            gbp = await asyncio.to_thread(analyze_gbp, gbp_url, gbp_competitor_urls or [])

        # Competitor gap analysis (from biz_data nearest_competitors reviews_text)
        competitor_gap = {"available": False}
        if biz_data and biz_data.get("nearest_competitors"):
            category = biz_data.get("category", "local business")
            competitor_gap = await asyncio.to_thread(
                analyze_competitor_gap, biz_data["nearest_competitors"], category
            )

        await send_progress(job_id, 92, "Calculating revenue impact…")
        revenue_impact = await asyncio.to_thread(
            calculate_revenue_impact,
            total_pages, scores,
            technical, onpage, schema, aeo, geo,
            performance, images, local_seo, conversion, content,
        )

        await send_progress(job_id, 93, "Generating AI recommendations & cold emails… (this may take a minute)")
        ai_recommendations = await asyncio.to_thread(
            generate_ai_recommendations,
            parsed_domain, total_pages, scores,
            technical, onpage, schema, aeo, geo,
            performance, images, local_seo, conversion, content,
        )

        await send_progress(job_id, 95, "Generating cold email drafts…")
        cold_emails = await asyncio.to_thread(
            generate_cold_emails,
            parsed_domain, scores, revenue_impact,
            competitors[0] if competitors else {}, wayback, local_seo,
        )

        await send_progress(job_id, 96, "Scoring lead quality…")
        lead_score = calculate_lead_score(
            scores, technical, onpage, performance,
            local_seo, conversion, revenue_impact,
            total_pages, gbp, keywords,
        )

        await send_progress(job_id, 96, "Building 90-day roadmap…")
        roadmap = await asyncio.to_thread(
            generate_roadmap,
            parsed_domain, scores, local_seo,
            trust_signals, competitor_gap, lead_score,
        )

        await send_progress(job_id, 96, "Generating PDF report…")
        report_path = await asyncio.to_thread(
            generate_report,
            job_id,
            url,
            total_pages,
            scores,
            technical,
            onpage,
            schema,
            aeo,
            geo,
            performance,
            images,
            local_seo,
            conversion,
            content,
            ai_recommendations,
            revenue_impact,
            wayback,
            competitors,
            cold_emails,
            progress_data,
            keywords,
            gbp,
            lead_score,
            trust_signals,
            mobile_screenshots,
            competitor_gap,
            roadmap,
            meta_rewrites,
        )

        await send_progress(job_id, 98, "Finalizing report…")

        # Build result payload (serializable data for WebSocket)
        result_data = {
            "job_id": job_id,
            "root_url": url,
            "total_pages": total_pages,
            "report_path": report_path,
            "scores": _make_serializable(scores),
            "technical": _make_serializable({
                "score": technical.get("score", 0),
                "summary": technical.get("summary", {}),
                "issues_4xx": technical.get("issues_4xx", [])[:20],
                "redirect_chains": technical.get("redirect_chains", [])[:10],
                "slow_pages": technical.get("slow_pages", [])[:10],
                "http_pages": technical.get("http_pages", [])[:10],
            }),
            "onpage": _make_serializable({
                "score": onpage.get("score", 0),
                "summary": onpage.get("summary", {}),
            }),
            "schema": _make_serializable({
                "score": schema.get("score", 0),
                "summary": schema.get("summary", {}),
                "all_schema_types_found": schema.get("all_schema_types_found", []),
                "missing_schema_types": schema.get("missing_schema_types", [])[:10],
                "has_about_page": False,
            }),
            "aeo": _make_serializable({
                "score": aeo.get("score", 0),
                "summary": aeo.get("summary", {}),
            }),
            "geo": _make_serializable({
                "score": geo.get("score", 0),
                "summary": geo.get("summary", {}),
                "has_about_page": geo.get("has_about_page", False),
                "has_contact_page": geo.get("has_contact_page", False),
                "brand_name": geo.get("brand_name", ""),
            }),
            "performance": _make_serializable({
                "score": performance.get("score", 0),
                "summary": performance.get("summary", {}),
            }),
            "images": _make_serializable({
                "score": images.get("score", 0),
                "summary": images.get("summary", {}),
            }),
            "local_seo": _make_serializable({
                "score": local_seo.get("score", 0),
                "summary": local_seo.get("summary", {}),
                "has_contact_page": local_seo.get("has_contact_page", False),
                "has_google_maps": local_seo.get("has_google_maps", False),
                "has_local_business_schema": local_seo.get("has_local_business_schema", False),
                "has_nap_info": local_seo.get("has_nap_info", False),
                "has_review_schema": local_seo.get("has_review_schema", False),
                "has_opening_hours": local_seo.get("has_opening_hours", False),
                "location_optimized_pages": local_seo.get("location_optimized_pages", 0),
                "missing_local_signals": local_seo.get("missing_local_signals", []),
            }),
            "conversion": _make_serializable({
                "score": conversion.get("score", 0),
                "summary": conversion.get("summary", {}),
                "pages_missing_cta": conversion.get("pages_missing_cta", [])[:15],
            }),
            "content": _make_serializable({
                "score": content.get("score", 0),
                "summary": content.get("summary", {}),
                "thin_content_pages": content.get("thin_content_pages", [])[:15],
            }),
            "revenue_impact": _make_serializable(revenue_impact),
            "wayback": _make_serializable(wayback),
            "competitors": _make_serializable(competitors),
            "cold_emails": _make_serializable(cold_emails),
            "progress": _make_serializable(progress_data),
            "keywords": _make_serializable({
                "score": keywords.get("score", 0),
                "summary": keywords.get("summary", {}),
                "opportunities": keywords.get("opportunities", [])[:20],
                "quick_wins": keywords.get("quick_wins", [])[:10],
                "top_keywords": keywords.get("top_keywords", [])[:30],
            }),
            "gbp": _make_serializable(gbp),
            "lead_score": _make_serializable(lead_score),
        }

        # Store full data in job
        jobs[job_id]["status"] = "complete"
        jobs[job_id]["data"] = result_data

        await send_progress(job_id, 100, "Analysis complete! Preparing results…")

        # Send complete message
        if job_id in ws_connections:
            try:
                await ws_connections[job_id].send_json({
                    "type": "complete",
                    "data": result_data,
                })
            except Exception:
                pass

    except Exception as e:
        error_msg = str(e)
        tb = traceback.format_exc()
        print(f"[Job {job_id}] Error: {error_msg}\n{tb}")
        jobs[job_id]["status"] = "error"
        jobs[job_id]["error"] = error_msg

        if job_id in ws_connections:
            try:
                await ws_connections[job_id].send_json({
                    "type": "error",
                    "message": f"Analysis failed: {error_msg}",
                })
            except Exception:
                pass


def _make_serializable(obj):
    """Recursively ensure all values are JSON-serializable."""
    if isinstance(obj, dict):
        return {k: _make_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, list):
        return [_make_serializable(item) for item in obj]
    elif isinstance(obj, (int, float, str, bool)) or obj is None:
        return obj
    else:
        return str(obj)


# ====== ENTRY POINT ======
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
